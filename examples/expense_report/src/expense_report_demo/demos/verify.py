"""Deterministic post-run verification primitives for demos.

Checks run before teardown (while the demo's artifacts are still in place) and
carry a failure taxonomy: `agent` (the work is wrong), `harness` (the check
itself couldn't observe the world, e.g. a missing automation grant). A passed
check has no category.
"""

from __future__ import annotations

import subprocess
from pathlib import Path
from typing import Literal

from openpyxl import load_workbook
from pydantic import BaseModel, ConfigDict, Field

VerifyFailureCategory = Literal["agent", "harness"]

_OSASCRIPT_TIMEOUT_S = 30.0
_AMOUNT_TOL = 0.01


def _row_is_populated(row: tuple[object, ...]) -> bool:
    """A row counts as populated if any cell holds non-blank content."""
    return any(cell is not None and str(cell).strip() != "" for cell in row)


def template_row_extent(sheet_path: Path) -> tuple[int, int]:
    """(count of populated rows, 1-based index of the last populated row) in the first worksheet."""
    resolved = sheet_path.expanduser()
    if not resolved.is_file():
        raise FileNotFoundError(f"sheet not found: {resolved}")
    wb = load_workbook(filename=resolved, data_only=True)
    sheet = wb.worksheets[0]
    count = 0
    last = 0
    for idx, row in enumerate(sheet.iter_rows(values_only=True), start=1):
        if _row_is_populated(row):
            count += 1
            last = idx
    return count, last


class VerifyCheck(BaseModel):
    """One named pass/fail observation about the post-run world."""

    model_config = ConfigDict(frozen=True, extra="forbid")
    name: str = Field(min_length=1)
    passed: bool
    reason: str = Field(min_length=1)
    failure_category: VerifyFailureCategory | None


def check_new_xlsx_rows_match_totals(
    *,
    name: str,
    sheet_path: Path,
    template_row_count: int,
    amount_column: int,
    expected_totals: list[float],
) -> VerifyCheck:
    """Rows appended past `template_row_count` must carry exactly `expected_totals`.

    Amounts compare as absolute values (ledgers book expenses negative) with a
    small tolerance, order-independent. Reads cached values (`data_only`), so
    an unsaved sheet reads as missing rows — saving is part of the task.
    """
    resolved = sheet_path.expanduser()
    if not resolved.is_file():
        return VerifyCheck(name=name, passed=False, reason=f"sheet not found: {resolved}", failure_category="harness")
    try:
        wb = load_workbook(filename=resolved, data_only=True)
    except (OSError, ValueError) as exc:
        return VerifyCheck(
            name=name, passed=False, reason=f"could not open {resolved}: {exc}", failure_category="harness"
        )

    sheet = wb.worksheets[0]
    new_rows = [
        row
        for idx, row in enumerate(sheet.iter_rows(values_only=True), start=1)
        if idx > template_row_count and _row_is_populated(row)
    ]
    if len(new_rows) != len(expected_totals):
        return VerifyCheck(
            name=name,
            passed=False,
            reason=f"expected {len(expected_totals)} new rows after row {template_row_count}, found {len(new_rows)}",
            failure_category="agent",
        )

    amounts: list[float] = []
    for row_idx, row in enumerate(new_rows, start=template_row_count + 1):
        value = row[amount_column] if amount_column < len(row) else None
        if not isinstance(value, (int, float)):
            return VerifyCheck(
                name=name,
                passed=False,
                reason=f"row {row_idx}: amount cell is not numeric ({value!r})",
                failure_category="agent",
            )
        amounts.append(abs(float(value)))

    unmatched = _unmatched_totals(amounts, expected_totals)
    if unmatched:
        return VerifyCheck(
            name=name,
            passed=False,
            reason=f"amounts not matching any expected receipt total: {unmatched}",
            failure_category="agent",
        )
    return VerifyCheck(
        name=name,
        passed=True,
        reason=f"{len(new_rows)} rows appended, all {len(expected_totals)} receipt totals present",
        failure_category=None,
    )


def _unmatched_totals(actual: list[float], expected: list[float]) -> list[float]:
    """Greedy 1:1 matching of sorted amount lists within tolerance; returns leftovers."""
    remaining = sorted(expected)
    leftovers: list[float] = []
    for amount in sorted(actual):
        for i, gold in enumerate(remaining):
            if abs(amount - gold) <= _AMOUNT_TOL:
                remaining.pop(i)
                break
        else:
            leftovers.append(amount)
    return leftovers + remaining


class DraftRef(BaseModel):
    """One Mail draft matching a (to_address, subject) query, keyed by its stable message id."""

    model_config = ConfigDict(frozen=True, extra="forbid")
    message_id: str = Field(min_length=1)
    attachment_count: int = Field(ge=0)


class MailScriptError(RuntimeError):
    """osascript against Mail failed or timed out — a harness condition, never the agent's fault."""


# Gmail (IMAP) drafts can't be reliably deleted from AppleScript — `delete`/`move` silently no-op
# because a Gmail draft is a label, not a folder item. So instead of clearing stale drafts, the
# caller snapshots existing draft ids before the run and the check counts only ids new since then
# (same baseline-diff shape as the ledger check). `message id` is assigned at creation and immutable.
def _list_matching_drafts(*, to_address: str, subject: str) -> list[DraftRef]:
    """Drafts to `to_address` with `subject`, as (message id, attachment count) pairs."""
    script = f"""
    tell application "Mail"
        set out to ""
        set matches to (every message of drafts mailbox whose subject is "{subject}")
        repeat with m in matches
            if (address of to recipients of m) contains "{to_address}" then
                set out to out & (message id of m) & tab & (count of mail attachments of m) & linefeed
            end if
        end repeat
        return out
    end tell
    """
    try:
        result = subprocess.run(
            ["osascript", "-e", script], check=False, capture_output=True, text=True, timeout=_OSASCRIPT_TIMEOUT_S
        )
    except subprocess.TimeoutExpired as exc:
        raise MailScriptError("Mail AppleScript timed out") from exc
    if result.returncode != 0:
        raise MailScriptError(f"Mail AppleScript failed ({result.returncode}): {result.stderr.strip()}")
    drafts: list[DraftRef] = []
    for line in result.stdout.splitlines():
        if not line.strip():
            continue
        message_id, _, count = line.partition("\t")
        drafts.append(DraftRef(message_id=message_id.strip(), attachment_count=int(count.strip())))
    return drafts


def snapshot_matching_draft_ids(*, to_address: str, subject: str) -> frozenset[str]:
    """Message ids of drafts already matching (to_address, subject), captured pre-run as a baseline.

    Raises MailScriptError if Mail can't be read, so a missing automation grant fails setup loudly
    rather than yielding an empty baseline that would let every stale draft read as the run's work.
    """
    return frozenset(d.message_id for d in _list_matching_drafts(to_address=to_address, subject=subject))


def evaluate_mail_draft(
    *,
    name: str,
    to_address: str,
    subject: str,
    drafts: list[DraftRef],
    baseline_message_ids: frozenset[str],
) -> VerifyCheck:
    """Pass iff a draft appeared since the baseline (id not in `baseline_message_ids`) with an attachment."""
    new_drafts = [d for d in drafts if d.message_id not in baseline_message_ids]
    if not new_drafts:
        return VerifyCheck(
            name=name,
            passed=False,
            reason=(
                f"no new draft to {to_address} with subject {subject!r} created this run "
                f"({len(baseline_message_ids)} pre-existing ignored)"
            ),
            failure_category="agent",
        )
    attachments = max(d.attachment_count for d in new_drafts)
    if attachments < 1:
        return VerifyCheck(
            name=name, passed=False, reason="new draft exists but has no attachment", failure_category="agent"
        )
    return VerifyCheck(
        name=name,
        passed=True,
        reason=f"new draft to {to_address} with {attachments} attachment(s)",
        failure_category=None,
    )


def check_mail_draft(*, name: str, to_address: str, subject: str, baseline_message_ids: frozenset[str]) -> VerifyCheck:
    """Read drafts via AppleScript, then grade against the pre-run baseline."""
    try:
        drafts = _list_matching_drafts(to_address=to_address, subject=subject)
    except MailScriptError as exc:
        return VerifyCheck(name=name, passed=False, reason=str(exc), failure_category="harness")
    return evaluate_mail_draft(
        name=name,
        to_address=to_address,
        subject=subject,
        drafts=drafts,
        baseline_message_ids=baseline_message_ids,
    )
